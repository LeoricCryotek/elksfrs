# -*- coding: utf-8 -*-
"""Budget CSV Import Wizard.

Allows lodge managers to import budget lines from a CSV file and optionally
roll forward a previous year's budget with a percentage adjustment.

CSV format:  AccountCode, AccountName, BudgetAmount, Note
"""
import base64
import csv
import io

from odoo import api, fields, models, _
from odoo.exceptions import UserError

import logging

_logger = logging.getLogger(__name__)

DEMO_CSV_HEADER = "AccountCode,AccountName,BudgetAmount,Note"
DEMO_CSV_ROWS = [
    "30205,Accounting,2000.00,Annual accountant retainer",
    "30270,Insurance G/L Property Plus,10500.00,5% increase from prior year",
    "30295,Janitorial Contract,5100.00,",
    "30363,Other Lodge Employees-Wages,29200.00,Projected with 5% raise",
    "40100,Liquor Sales,60000.00,Income projection",
    "50205,Food Purchases,19800.00,Based on prior year + 5%",
]


class BudgetImportWizard(models.TransientModel):
    """Wizard to import budget lines from CSV or roll forward a prior budget."""

    _name = "budget.import.wizard"
    _description = "Budget Import Wizard"

    import_mode = fields.Selection([
        ('csv', 'Import from CSV / Excel File'),
        ('qb_pnl', 'QuickBooks Import Update (XLSX)'),
        ('rollforward', 'Roll Forward Prior Year Budget'),
    ], string="Import Mode", default='csv', required=True)

    # --- CSV / QB P&L import fields ---
    file_data = fields.Binary("File")
    file_name = fields.Char("Filename")

    # --- QB P&L options ---
    skip_dues_accounts = fields.Boolean(
        "Skip Member Dues Lines", default=True,
        help="Skip member dues accounts (30010, 30011, 30020, 30021). "
             "Dues income should come from CLMS import or manual entry only.",
    )

    # --- Roll-forward fields ---
    source_budget_id = fields.Many2one(
        "elks.budget", string="Source Budget",
        help="The prior-year budget to copy lines from.",
        domain="[('state', 'in', ('board_approved', 'floor_approved', 'submitted'))]",
    )
    adjustment_pct = fields.Float(
        "Adjustment %", default=5.0,
        help="Percentage to increase (positive) or decrease (negative) "
             "all expense lines.  Income lines are copied as-is.",
    )

    # --- Common fields ---
    target_budget_id = fields.Many2one(
        "elks.budget", string="Target Budget", required=True,
        domain="[('state', '=', 'draft')]",
        help="The draft budget to import lines into.",
    )
    overwrite = fields.Boolean(
        "Replace Existing Lines", default=False,
        help="If checked, existing budget lines for matching accounts "
             "will be updated.  Otherwise, duplicates are skipped.",
    )
    create_missing_accounts = fields.Boolean(
        "Create Missing Accounts", default=False,
        help="Automatically create Elks accounts for codes not yet "
             "in the Chart of Accounts.",
    )

    # --- Results ---
    state = fields.Selection([
        ('setup', 'Setup'),
        ('done', 'Done'),
    ], default='setup')
    result_message = fields.Text("Import Results", readonly=True)

    # --- Demo CSV download ---
    demo_csv = fields.Binary("Demo CSV", readonly=True)
    demo_csv_name = fields.Char(default="budget_import_template.csv")

    def action_download_demo(self):
        """Generate and return a blank demo CSV template."""
        self.ensure_one()
        output = io.StringIO()
        output.write(DEMO_CSV_HEADER + "\n")
        for row in DEMO_CSV_ROWS:
            output.write(row + "\n")

        self.demo_csv = base64.b64encode(output.getvalue().encode("utf-8"))
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import(self):
        """Dispatch to CSV import or roll-forward based on mode."""
        self.ensure_one()
        if not self.target_budget_id:
            raise UserError(_("Please select a target budget."))
        if self.target_budget_id.state != 'draft':
            raise UserError(_("You can only import into a Draft budget."))

        if self.import_mode == 'csv':
            result = self._import_csv()
        elif self.import_mode == 'qb_pnl':
            result = self._import_qb_pnl()
        else:
            result = self._rollforward()

        self.write({
            'state': 'done',
            'result_message': result,
        })
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    # ------------------------------------------------------------------
    # CSV Import
    # ------------------------------------------------------------------

    def _import_csv(self):
        if not self.file_data:
            raise UserError(_("Please upload a CSV or Excel file."))

        raw = base64.b64decode(self.file_data)
        fname = (self.file_name or '').lower()

        # Detect Excel files by extension or magic bytes
        is_xlsx = fname.endswith(('.xlsx', '.xls')) or raw[:4] == b'PK\x03\x04'
        if is_xlsx:
            return self._import_xlsx(raw)

        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw.decode('latin-1')

        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            raise UserError(_("Empty or invalid CSV file."))

        # Flexible column matching
        code_col = self._find_col(reader.fieldnames, ['accountcode', 'account', 'code', 'acctcode'])
        name_col = self._find_col(reader.fieldnames, ['accountname', 'name', 'description'])
        amount_col = self._find_col(reader.fieldnames, ['budgetamount', 'amount', 'budget', 'annual'])
        note_col = self._find_col(reader.fieldnames, ['note', 'memo', 'comment'])

        if not code_col:
            raise UserError(_(
                "Could not find an account code column.  "
                "Expected: AccountCode, Account, or Code.  "
                "Found: %s"
            ) % ", ".join(reader.fieldnames))
        if not amount_col:
            raise UserError(_(
                "Could not find a budget amount column.  "
                "Expected: BudgetAmount, Amount, or Budget.  "
                "Found: %s"
            ) % ", ".join(reader.fieldnames))

        created = 0
        created_accounts = 0
        updated = 0
        skipped_empty = 0
        skipped_not_found = []
        skipped_duplicate = []
        errors = []
        created_lines = []

        for i, row in enumerate(reader, start=2):
            try:
                acct_code = str(row.get(code_col, '')).strip()
                acct_name = str(row.get(name_col, '')).strip() if name_col else ''
                amount_str = str(row.get(amount_col, '')).strip()
                note = str(row.get(note_col, '')).strip() if note_col else ''

                if not acct_code:
                    skipped_empty += 1
                    continue

                amount = self._parse_amount(amount_str)
                acct_existed = bool(self.env['elks.account'].search([
                    '|',
                    ('code', '=', acct_code),
                    '&', ('code', '=', acct_code[:5]),
                         ('subaccount', '=', acct_code[5:] if len(acct_code) > 5 else False),
                ], limit=1))

                acct = self._find_or_create_account(acct_code, acct_name)
                if not acct:
                    skipped_not_found.append(
                        f"  Row {i}: {acct_code} — {acct_name}"
                    )
                    continue

                if not acct_existed:
                    created_accounts += 1

                result = self._upsert_budget_line(acct, amount, note)
                if result == 'created':
                    created += 1
                    created_lines.append(f"  {acct_code} {acct_name}: ${amount:,.2f}")
                elif result == 'updated':
                    updated += 1
                else:
                    skipped_duplicate.append(
                        f"  Row {i}: {acct_code} — {acct_name} "
                        f"(line already exists, 'Replace Existing' not checked)"
                    )

            except Exception as e:
                errors.append(f"  Row {i}: {acct_code if 'acct_code' in dir() else '?'} — {e}")

        # Build detailed results message
        parts = []
        parts.append(f"IMPORT RESULTS: {created} created, {updated} updated")
        total_skipped = skipped_empty + len(skipped_not_found) + len(skipped_duplicate)
        if total_skipped:
            parts[0] += f", {total_skipped} skipped"
        if created_accounts:
            parts.append(f"\nNew accounts created in Chart of Accounts: {created_accounts}")

        if skipped_not_found:
            parts.append(f"\n--- SKIPPED: Account not found ({len(skipped_not_found)}) ---")
            parts.append("Enable 'Create Missing Accounts' to auto-create these:")
            parts.extend(skipped_not_found)

        if skipped_duplicate:
            parts.append(f"\n--- SKIPPED: Duplicate lines ({len(skipped_duplicate)}) ---")
            parts.append("Check 'Replace Existing Lines' to update these instead:")
            parts.extend(skipped_duplicate)

        if skipped_empty:
            parts.append(f"\n--- SKIPPED: Empty account code ({skipped_empty} rows) ---")

        if errors:
            parts.append(f"\n--- ERRORS ({len(errors)}) ---")
            parts.extend(errors)

        return "\n".join(parts)

    # ------------------------------------------------------------------
    # Excel Budget Import (XLSX — "Budget vs Actual" export format)
    # ------------------------------------------------------------------

    def _import_xlsx(self, raw):
        """Import budget lines from an Excel file.

        Supports the "Budget vs Actual" export format where:
        - Row 5 has headers: Account Code, Account Name, Budgeted, ...
        - Data rows start at row 6
        - Skips section headers (INCOME, EXPENSES, TOTAL, NET, etc.)
        """
        import openpyxl

        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:
            raise UserError(_(
                "Could not open file as Excel: %s\n"
                "Make sure the file is a valid .xlsx file."
            ) % e)

        ws = wb.active

        # Find header row — look for "Account" in col A or B
        header_row = None
        code_col = name_col = amount_col = note_col = None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, min(ws.max_column + 1, 10)):
                val = str(ws.cell(r, c).value or '').lower().strip()
                if 'account' in val and 'code' in val:
                    header_row = r
                    code_col = c
                elif val in ('account name', 'name', 'description'):
                    if header_row is None:
                        header_row = r
                    name_col = c
                elif val in ('budgeted', 'budget', 'amount', 'budgetamount'):
                    amount_col = c
                elif val in ('note', 'memo'):
                    note_col = c
            if header_row == r:
                break

        if not header_row or not code_col:
            # Fallback: assume col A = code, B = name, C = amount
            header_row = 5
            code_col = 1
            name_col = 2
            amount_col = 3

        if not amount_col:
            amount_col = 3

        skip_labels = {
            'income', 'expenses', 'expense', 'total', 'net',
            'none', '', 'total income', 'total expenses',
        }

        created = 0
        created_accounts = 0
        updated = 0
        skipped_empty = 0
        skipped_not_found = []
        skipped_duplicate = []
        errors = []
        created_lines = []

        name_mismatches = []

        for r in range(header_row + 1, ws.max_row + 1):
            try:
                acct_code = str(ws.cell(r, code_col).value or '').strip()
                acct_name = str(ws.cell(r, name_col).value or '').strip() if name_col else ''
                amount_raw = ws.cell(r, amount_col).value if amount_col else None

                if not acct_code or acct_code.lower() in skip_labels:
                    skipped_empty += 1
                    continue

                # Skip TOTAL / NET rows
                if acct_code.upper().startswith('TOTAL') or acct_code.upper().startswith('NET'):
                    skipped_empty += 1
                    continue

                # Parse amount
                if amount_raw is None:
                    amount = 0.0
                elif isinstance(amount_raw, (int, float)):
                    amount = float(amount_raw)
                else:
                    amount = self._parse_amount(str(amount_raw))

                acct_existed = bool(self.env['elks.account'].search([
                    '|',
                    ('code', '=', acct_code),
                    '&', ('code', '=', acct_code[:5]),
                         ('subaccount', '=', acct_code[5:] if len(acct_code) > 5 else False),
                ], limit=1))

                acct = self._find_or_create_account(acct_code, acct_name)
                if not acct:
                    skipped_not_found.append(
                        f"  Row {r}: {acct_code} — {acct_name}"
                    )
                    continue

                if not acct_existed:
                    created_accounts += 1

                # Log every account for traceability
                _logger.info(
                    "Budget XLSX import row %d: code=%s  "
                    "excel_name='%s'  db_name='%s'  db_id=%d  amount=%.2f",
                    r, acct_code, acct_name, acct.name, acct.id, amount,
                )

                # Check for name mismatch and auto-fix
                if acct_name and acct.name and acct_name != acct.name:
                    old_name = acct.name
                    acct.write({'name': acct_name})
                    name_mismatches.append(
                        f"  {acct_code}: '{old_name}' → '{acct_name}' (FIXED)"
                    )
                    _logger.info(
                        "Budget XLSX import: FIXED account %s name "
                        "'%s' → '%s' (id=%d)",
                        acct_code, old_name, acct_name, acct.id,
                    )

                note = ''
                if note_col:
                    note = str(ws.cell(r, note_col).value or '').strip()

                result = self._upsert_budget_line(acct, amount, note)
                if result == 'created':
                    created += 1
                    created_lines.append(
                        f"  {acct_code} {acct_name}: ${amount:,.2f}"
                    )
                elif result == 'updated':
                    updated += 1
                else:
                    skipped_duplicate.append(
                        f"  Row {r}: {acct_code} — {acct_name} "
                        f"(line already exists, 'Replace Existing' not checked)"
                    )

            except Exception as e:
                errors.append(
                    f"  Row {r}: {acct_code if 'acct_code' in dir() else '?'} — {e}"
                )

        # Build results
        parts = []
        parts.append(f"EXCEL IMPORT: {created} created, {updated} updated")
        total_skipped = skipped_empty + len(skipped_not_found) + len(skipped_duplicate)
        if total_skipped:
            parts[0] += f", {total_skipped} skipped"
        if created_accounts:
            parts.append(
                f"\nNew accounts created in Chart of Accounts: {created_accounts}"
            )

        if name_mismatches:
            parts.append(
                f"\n--- ACCOUNT NAMES CORRECTED ({len(name_mismatches)}) ---"
            )
            parts.append(
                "These account names were updated to match the Excel file:"
            )
            parts.extend(name_mismatches)

        if skipped_not_found:
            parts.append(
                f"\n--- SKIPPED: Account not found ({len(skipped_not_found)}) ---"
            )
            parts.append(
                "Enable 'Create Missing Accounts' to auto-create these:"
            )
            parts.extend(skipped_not_found)

        if skipped_duplicate:
            parts.append(
                f"\n--- SKIPPED: Duplicate lines ({len(skipped_duplicate)}) ---"
            )
            parts.append(
                "Check 'Replace Existing Lines' to update these instead:"
            )
            parts.extend(skipped_duplicate)

        if errors:
            parts.append(f"\n--- ERRORS ({len(errors)}) ---")
            parts.extend(errors)

        return "\n".join(parts)

    # ------------------------------------------------------------------
    # QuickBooks P&L Import (XLSX)
    # ------------------------------------------------------------------

    # Member dues account codes — these should only be updated via
    # CLMS import or manual entry, never from a QB P&L import.
    DUES_ACCOUNT_CODES = {'30010', '30011', '30020', '30021'}

    def _import_qb_pnl(self):
        """Import a QuickBooks Profit & Loss report (XLSX) into budget lines.

        The QB P&L XLSX has a hierarchical layout:
          Column E: main accounts  ("30205 · Accounting")
          Column F: sub-accounts   ("30320A · Supplies")
          Column G: sub-sub-accts  ("90205C1 · Summer Youth Program")
          Column H: amounts

        "Total ..." rows are skipped.  "... - Other" lines map to the
        base account (no subaccount suffix).
        """
        if not self.file_data:
            raise UserError(_("Please upload a QuickBooks P&L file (.xlsx)."))

        import openpyxl

        raw = base64.b64decode(self.file_data)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as e:
            raise UserError(_(
                "Could not open file as XLSX: %s\n"
                "Make sure you exported the P&L from QuickBooks as an Excel file."
            ) % e)

        ws = wb.active

        created = 0
        created_accounts = 0
        updated = 0
        skipped_dues = []
        skipped_not_found = []
        skipped_duplicate = []
        skipped_total = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), start=1):
            try:
                # Find the account label — check columns E(5), F(6), G(7)
                acct_label = None
                for col_idx in (5, 6, 7):  # E, F, G
                    cell = row[col_idx - 1] if col_idx <= len(row) else None
                    if cell and cell.value and str(cell.value).strip():
                        acct_label = str(cell.value).strip()
                        break

                if not acct_label:
                    continue

                # Skip "Total ..." rows
                if acct_label.lower().startswith('total '):
                    skipped_total += 1
                    continue

                # Get the amount from column H (8)
                amt_cell = row[7] if len(row) >= 8 else None
                if not amt_cell or amt_cell.value is None:
                    continue
                try:
                    amount = float(amt_cell.value)
                except (ValueError, TypeError):
                    continue

                # Extract account code and name from "CODE · Name" format
                acct_code, acct_name = self._extract_qb_account(acct_label)
                if not acct_code:
                    continue

                # Strip "- Other" suffix from name (these are the base account catch-all)
                is_other = False
                if acct_name and acct_name.endswith('- Other'):
                    acct_name = acct_name.replace('- Other', '').strip()
                    is_other = True

                # Convert letter subaccount suffixes to numeric
                # e.g. "30320A" → "3032001", "90205C1" → "9020531"
                if len(acct_code) > 5 and not is_other:
                    base_code = acct_code[:5]
                    letter_sub = acct_code[5:]
                    num_sub = self._letter_sub_to_num(letter_sub)
                    acct_code = base_code + num_sub
                else:
                    base_code = acct_code

                # Skip dues accounts if requested
                if self.skip_dues_accounts and base_code in self.DUES_ACCOUNT_CODES:
                    skipped_dues.append(
                        f"  {acct_code} — {acct_name}: ${amount:,.2f}"
                    )
                    continue

                # Find or create account
                acct_existed = bool(self._find_existing_account(acct_code))
                acct = self._find_or_create_account(acct_code, acct_name)
                if not acct:
                    skipped_not_found.append(
                        f"  Row {row_idx}: {acct_code} — {acct_name}"
                    )
                    continue

                # Update account name if it was auto-created with a generic
                # name or if the QB name differs (QB is authoritative)
                if acct_name and acct.name != acct_name:
                    acct.write({'name': acct_name})

                if not acct_existed:
                    created_accounts += 1

                result = self._upsert_budget_line(acct, amount, f"QB P&L: {acct_name}")
                if result == 'created':
                    created += 1
                elif result == 'updated':
                    updated += 1
                else:
                    skipped_duplicate.append(
                        f"  Row {row_idx}: {acct_code} — {acct_name}"
                    )

            except Exception as e:
                errors.append(f"  Row {row_idx}: {e}")

        # Build results message
        parts = [f"QB P&L IMPORT: {created} created, {updated} updated"]
        total_skip = len(skipped_dues) + len(skipped_not_found) + len(skipped_duplicate)
        if total_skip:
            parts[0] += f", {total_skip} skipped"
        if created_accounts:
            parts.append(f"\nNew accounts created: {created_accounts}")

        if skipped_dues:
            parts.append(f"\n--- SKIPPED: Member Dues ({len(skipped_dues)}) ---")
            parts.append("These are updated via CLMS import or manual entry:")
            parts.extend(skipped_dues)

        if skipped_not_found:
            parts.append(f"\n--- SKIPPED: Account not found ({len(skipped_not_found)}) ---")
            parts.append("Enable 'Create Missing Accounts' to auto-create these:")
            parts.extend(skipped_not_found)

        if skipped_duplicate:
            parts.append(f"\n--- SKIPPED: Duplicate lines ({len(skipped_duplicate)}) ---")
            parts.append("Check 'Replace Existing Lines' to update these instead:")
            parts.extend(skipped_duplicate)

        if errors:
            parts.append(f"\n--- ERRORS ({len(errors)}) ---")
            parts.extend(errors)

        return "\n".join(parts)

    def _find_existing_account(self, code):
        """Check if an account exists without creating it."""
        Account = self.env['elks.account']
        acct = Account.search([('code', '=', code)], limit=1)
        if acct:
            return acct
        if len(code) > 5:
            return Account.search([
                ('code', '=', code[:5]),
                ('subaccount', '=', code[5:]),
            ], limit=1)
        return False

    @staticmethod
    def _letter_sub_to_num(sub):
        """Convert letter-based subaccount to numeric.

        QuickBooks uses letters (A, B, C1, etc.) while the Elks COA
        uses numeric subaccounts (01, 02, 31, etc.).
          A → 01, B → 02, ..., Z → 26
          C1 → 31  (C=3, 1 appended)
        """
        result = ''
        for ch in sub:
            if ch.isalpha():
                result += str(ord(ch.upper()) - ord('A') + 1)
            else:
                result += ch
        return result.zfill(2) if len(result) < 2 else result[:2]

    @staticmethod
    def _extract_qb_account(label):
        """Extract account code and name from QB format.

        Handles:
          "30205 · Accounting"        → ("30205", "Accounting")
          "30320A · Supplies"         → ("30320A", "Supplies")
          "90205C1 · Summer Youth"    → ("90205C1", "Summer Youth")
          "30600 · Fund Raiser - Other" → ("30600", "Fund Raiser - Other")
          "Income"                    → (None, "Income")
        """
        if not label:
            return None, ''

        # Split on " · " (QB uses middle dot separator)
        for sep in (' · ', ' · ', ' \u00b7 '):
            if sep in label:
                parts = label.split(sep, 1)
                code_part = parts[0].strip()
                name_part = parts[1].strip() if len(parts) > 1 else ''
                if code_part and code_part[0].isdigit():
                    return code_part, name_part
                return None, label

        # No separator — check if starts with digits
        if label and label[0].isdigit():
            parts = label.split(None, 1)
            return parts[0], (parts[1] if len(parts) > 1 else '')

        return None, label

    # ------------------------------------------------------------------
    # Roll Forward
    # ------------------------------------------------------------------

    def _rollforward(self):
        if not self.source_budget_id:
            raise UserError(_("Please select a source budget to roll forward."))

        pct = self.adjustment_pct / 100.0
        created = 0
        updated = 0
        skipped = 0

        for src_line in self.source_budget_id.line_ids:
            if not src_line.account_id:
                continue

            # Income lines copied as-is; expense/cogs lines get the adjustment
            if src_line.account_type in ('expense', 'cogs'):
                new_amount = round(src_line.amount * (1.0 + pct), 2)
            else:
                new_amount = src_line.amount

            note = f"Rolled forward from {self.source_budget_id.name}"
            if src_line.account_type in ('expense', 'cogs') and pct != 0:
                note += f" (+{self.adjustment_pct:.1f}%)"

            result = self._upsert_budget_line(src_line.account_id, new_amount, note)
            if result == 'created':
                created += 1
            elif result == 'updated':
                updated += 1
            else:
                skipped += 1

        return (
            f"Roll-forward complete from {self.source_budget_id.name}: "
            f"{created} created, {updated} updated, {skipped} skipped."
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _upsert_budget_line(self, account, amount, note=''):
        """Create or update a budget line for the given account."""
        BudgetLine = self.env['elks.budget.line']
        existing = BudgetLine.search([
            ('budget_id', '=', self.target_budget_id.id),
            ('account_id', '=', account.id),
        ], limit=1)

        if existing:
            if self.overwrite:
                existing.write({'amount': amount, 'note': note})
                return 'updated'
            return 'skipped'

        BudgetLine.create({
            'budget_id': self.target_budget_id.id,
            'account_id': account.id,
            'amount': amount,
            'note': note,
        })
        return 'created'

    def _find_or_create_account(self, code, name=''):
        """Locate an Elks account by code, optionally creating it.

        Also corrects the account_type if the code range indicates a
        different type than what is currently stored (e.g. an income
        account that was mistakenly created as expense).
        """
        Account = self.env['elks.account']
        # Try exact match first
        acct = Account.search([('code', '=', code)], limit=1)
        if acct:
            self._fix_account_type(acct)
            return acct

        # Try with subaccount (e.g. "30320A" → code=30320, sub=A)
        if len(code) > 5:
            base = code[:5]
            sub = code[5:]
            acct = Account.search([
                ('code', '=', base),
                ('subaccount', '=', sub),
            ], limit=1)
            if acct:
                self._fix_account_type(acct)
                return acct

        if not self.create_missing_accounts:
            return False

        # Determine base code and optional subaccount for creation
        base_code = code
        sub = ''
        if len(code) > 5:
            base_code = code[:5]
            sub = code[5:]

        acct_type = self._guess_account_type(base_code)

        # Find parent account for subaccounts
        parent = False
        if sub:
            parent_acct = Account.search([('code', '=', base_code), ('subaccount', '=', False)], limit=1)
            if parent_acct:
                parent = parent_acct.id

        vals = {
            'code': base_code,
            'name': name or f"Imported {code}",
            'account_type': acct_type,
        }
        if sub:
            vals['subaccount'] = sub
        if parent:
            vals['parent_id'] = parent
            # Inherit department from parent
            parent_acct = Account.browse(parent)
            if parent_acct.department_id:
                vals['department_id'] = parent_acct.department_id.id

        # Auto-assign department from code range if not inherited
        if 'department_id' not in vals:
            dept = self._guess_department(base_code)
            if dept:
                vals['department_id'] = dept.id

        return Account.create(vals)

    def _guess_department(self, code):
        """Return the elks.department matching an account code range."""
        if not code:
            return False
        prefix = code[:2] if len(code) >= 2 else code
        dept_map = {
            '10': '10', '15': '10', '20': '10', '21': '10',
            '23': '10', '29': '10',
            '30': '30', '40': '40', '50': '50', '60': '60',
            '61': '61', '62': '62', '63': '63', '64': '64',
            '65': '65', '66': '66', '67': '67',
            '90': '90', '91': '90', '92': '90', '93': '90',
            '94': '90', '95': '90',
        }
        dept_code = dept_map.get(prefix)
        if dept_code:
            return self.env['elks.department'].search(
                [('code', '=', dept_code)], limit=1,
            )
        return False

    def _fix_account_type(self, account):
        """Correct account_type and fill missing department."""
        updates = {}
        expected = self._guess_account_type(account.code)
        if expected and account.account_type != expected:
            _logger.info(
                "Correcting account %s type: %s → %s",
                account.code, account.account_type, expected,
            )
            updates['account_type'] = expected
        if not account.department_id:
            dept = self._guess_department(account.code)
            if dept:
                updates['department_id'] = dept.id
        if updates:
            account.write(updates)

    @staticmethod
    def _find_col(fieldnames, candidates):
        """Find a column name matching any candidate (case-insensitive, no spaces)."""
        normalized = {h.lower().replace(' ', '').replace('_', ''): h for h in fieldnames}
        for c in candidates:
            if c in normalized:
                return normalized[c]
        return None

    @staticmethod
    def _guess_account_type(code):
        """Determine account type from the Elks Chart of Accounts code ranges.

        Lewiston Lodge #896 COA structure (QuickBooks P&L mapping):
          30010–30199  = Income (dues, rents, fees, misc income)
          30200–39999  = Expense (lodge operating expenses)
          40100–40199  = Income (lounge/bar sales)
          40200–40229  = COGS  (liquor, beer, wine, soda, sundries purchases)
          40230–49999  = Expense (bar operating expenses: wages, insurance, etc.)
          50100–50119  = Income (food service sales)
          50120–50209  = COGS  (food purchases)
          50210–59999  = Expense (dining operating expenses: supplies, wages, etc.)
          60000–65999  = Expense (general & admin)
          66100        = Income (RV park income)
          66200–67999  = Expense (RV operating expenses)
          90000–91999  = Income (donations, grants, fundraising)
          92000–95999  = Expense (community/charity expenses)
        Balance sheet prefixes:
          10xxx–12xxx  = Asset
          15xxx–16xxx  = Fixed Asset
          20xxx–21xxx  = Liability
          23xxx        = Long-term Liability
          29xxx        = Equity
        """
        try:
            num = int(code[:5].ljust(5, '0'))
        except (ValueError, TypeError):
            return 'expense'

        # Balance sheet ranges
        if 10000 <= num <= 12999:
            return 'asset'
        if 15000 <= num <= 16999:
            return 'fixed_asset'
        if 20000 <= num <= 21999:
            return 'liability'
        if 23000 <= num <= 23999:
            return 'long_term_liability'
        if 29000 <= num <= 29999:
            return 'equity'

        # P&L income ranges
        if 30010 <= num <= 30199:
            return 'income'
        if 40100 <= num <= 40199:
            return 'income'
        if 50100 <= num <= 50119:
            return 'income'
        if num == 66100:
            return 'income'
        if 90000 <= num <= 91999:
            return 'income'

        # P&L COGS ranges (purchases only)
        if 40200 <= num <= 40229:
            return 'cogs'
        if 50120 <= num <= 50209:
            return 'cogs'

        # P&L expense ranges (everything else)
        if 30200 <= num <= 39999:
            return 'expense'
        if 40230 <= num <= 49999:
            return 'expense'
        if 50210 <= num <= 59999:
            return 'expense'
        if 60000 <= num <= 67999:
            return 'expense'
        if 92000 <= num <= 95999:
            return 'expense'

        return 'expense'

    @staticmethod
    def _parse_amount(val):
        if not val:
            return 0.0
        val = str(val).strip().replace('$', '').replace(',', '').replace(' ', '')
        if val.startswith('(') and val.endswith(')'):
            val = '-' + val[1:-1]
        try:
            return round(float(val), 2)
        except ValueError:
            return 0.0
