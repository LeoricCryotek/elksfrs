# -*- coding: utf-8 -*-
"""Elks Lodge Budget Management.

Per the AA Manual, each Department (Class) budget must create sufficient
Income to cover Expenses and show an annual profit.  The overall budget
including all activities cannot be negative.

FRS budget CSV format: LodgeNumber, LodgeGLAccount, FYE, Version, Annual
"""
import base64
import csv
import datetime
import io
import logging

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


def _get_lodge_year_choices(self):
    """Generate lodge year selections for ±5 years around now."""
    current_year = datetime.date.today().year
    choices = []
    for yr in range(current_year - 3, current_year + 6):
        label = f"{yr}-{yr + 1}"
        choices.append((label, label))
    return choices


class ElksBudget(models.Model):
    _name = "elks.budget"
    _description = "Elks Lodge Annual Budget"
    _order = "fiscal_year_end desc"
    _inherit = ["mail.thread"]

    name = fields.Char(
        "Budget Name", compute="_compute_name", store=True,
    )
    fiscal_year_end = fields.Date(
        "Fiscal Year End (March 31)", required=True, index=True,
        help="March 31 of the year the lodge year ends.",
    )
    lodge_year = fields.Selection(
        selection=_get_lodge_year_choices,
        string="Lodge Year",
        help="Select the lodge year (April–March). The fiscal year end will auto-populate.",
    )
    state = fields.Selection([
        ('draft', 'Draft'),
        ('board_pending', 'Board Review'),
        ('board_approved', 'Board Approved'),
        ('floor_pending', 'Floor Review'),
        ('floor_approved', 'Floor Approved'),
        ('rejected', 'Rejected'),
        ('submitted', 'Submitted to FRS'),
    ], default='draft', tracking=True, index=True)

    line_ids = fields.One2many(
        "elks.budget.line", "budget_id", string="Budget Lines",
        copy=True,
    )
    expense_line_ids = fields.One2many(
        "elks.budget.line", "budget_id", string="Expense Lines",
        domain=[('account_type', 'in', ('expense', 'cogs'))],
    )
    income_line_ids = fields.One2many(
        "elks.budget.line", "budget_id", string="Income Lines",
        domain=[('account_type', '=', 'income')],
    )

    # Related amendment / transfer records
    transfer_ids = fields.One2many(
        "elks.budget.transfer", "budget_id", string="Transfers",
    )
    amendment_ids = fields.One2many(
        "elks.budget.amendment", "budget_id", string="Amendments",
    )
    transfer_count = fields.Integer(
        compute="_compute_amendment_counts",
    )
    amendment_count = fields.Integer(
        compute="_compute_amendment_counts",
    )

    total_income = fields.Monetary(
        "Total Income", compute="_compute_totals", store=True,
        currency_field='currency_id',
    )
    total_expense = fields.Monetary(
        "Total Expenses", compute="_compute_totals", store=True,
        currency_field='currency_id',
    )
    net_budget = fields.Monetary(
        "Net Budget", compute="_compute_totals", store=True,
        currency_field='currency_id',
    )

    # --- Actuals roll-up ---
    actual_income = fields.Monetary(
        "Actual Income", compute="_compute_actual_totals",
        currency_field='currency_id',
    )
    actual_expense = fields.Monetary(
        "Actual Expenses", compute="_compute_actual_totals",
        currency_field='currency_id',
    )
    actual_net = fields.Monetary(
        "Actual Net", compute="_compute_actual_totals",
        currency_field='currency_id',
    )
    income_variance = fields.Monetary(
        "Income Variance", compute="_compute_actual_totals",
        currency_field='currency_id',
        help="Actual income − budgeted income (positive = ahead of plan).",
    )
    expense_variance = fields.Monetary(
        "Expense Variance", compute="_compute_actual_totals",
        currency_field='currency_id',
        help="Budgeted expense − actual expense (positive = under budget).",
    )
    currency_id = fields.Many2one(
        "res.currency", default=lambda self: self.env.company.currency_id,
    )

    # Dashboard HTML bar chart — department budget vs actual
    department_chart_html = fields.Html(
        "Department Chart", compute="_compute_department_chart",
        sanitize=False,
    )

    # P&L statement HTML — grouped like QuickBooks
    pnl_html = fields.Html(
        "Profit & Loss", compute="_compute_pnl_html",
        sanitize=False,
    )

    # CSV export
    csv_file = fields.Binary("Budget CSV", readonly=True)
    csv_filename = fields.Char("CSV Filename")

    version = fields.Char(
        "Version", default="1",
        help="FRS budget version identifier.",
    )

    def _compute_amendment_counts(self):
        for rec in self:
            rec.transfer_count = len(rec.transfer_ids)
            rec.amendment_count = len(rec.amendment_ids)

    @api.depends("fiscal_year_end")
    def _compute_name(self):
        for rec in self:
            if rec.fiscal_year_end:
                start_yr = rec.fiscal_year_end.year - 1
                rec.name = f"Budget {start_yr}-{rec.fiscal_year_end.year}"
            else:
                rec.name = "New Budget"

    @api.onchange("lodge_year")
    def _onchange_lodge_year(self):
        """Auto-populate fiscal year end when a lodge year is selected."""
        if self.lodge_year:
            try:
                end_year = int(self.lodge_year.split("-")[1])
                self.fiscal_year_end = datetime.date(end_year, 3, 31)
            except (ValueError, IndexError):
                pass

    @api.depends("line_ids.amount", "line_ids.account_id.account_type")
    def _compute_totals(self):
        for rec in self:
            income = sum(
                l.amount for l in rec.line_ids
                if l.account_id and l.account_id.account_type == 'income'
            )
            expense = sum(
                l.amount for l in rec.line_ids
                if l.account_id and l.account_id.account_type in (
                    'expense', 'cogs'
                )
            )
            rec.total_income = income
            rec.total_expense = expense
            rec.net_budget = income - expense

    @api.depends("line_ids.actual_amount", "line_ids.account_type",
                 "total_income", "total_expense")
    def _compute_actual_totals(self):
        """Compute actual income/expense totals and variances from budget lines."""
        for rec in self:
            income_lines = rec.line_ids.filtered(
                lambda l: l.account_type == 'income'
            )
            expense_lines = rec.line_ids.filtered(
                lambda l: l.account_type in ('expense', 'cogs')
            )
            rec.actual_income = sum(income_lines.mapped('actual_amount'))
            rec.actual_expense = sum(expense_lines.mapped('actual_amount'))
            rec.actual_net = rec.actual_income - rec.actual_expense
            rec.income_variance = rec.actual_income - rec.total_income
            rec.expense_variance = rec.total_expense - rec.actual_expense

    def _compute_department_chart(self):
        """Build an HTML horizontal bar chart showing budget vs actual by dept."""
        for rec in self:
            if not rec.line_ids:
                rec.department_chart_html = (
                    '<div class="text-muted text-center p-4">'
                    '<i class="fa fa-bar-chart fa-2x mb-2"/><br/>'
                    'No budget lines yet — add lines to see the chart.</div>'
                )
                continue

            # Aggregate by department
            dept_data = {}
            for line in rec.line_ids:
                dept = line.department_id.name or 'Unassigned'
                if dept not in dept_data:
                    dept_data[dept] = {
                        'budgeted': 0.0, 'actual': 0.0,
                        'type': line.account_type,
                    }
                dept_data[dept]['budgeted'] += line.amount or 0.0
                dept_data[dept]['actual'] += line.actual_amount or 0.0

            # Sort by budgeted descending
            sorted_depts = sorted(
                dept_data.items(), key=lambda x: abs(x[1]['budgeted']), reverse=True,
            )
            max_val = max(
                max(abs(d['budgeted']), abs(d['actual'])) for _, d in sorted_depts
            ) or 1.0

            rows = []
            for dept_name, d in sorted_depts:
                budgeted = d['budgeted']
                actual = d['actual']
                bud_pct = min(abs(budgeted) / max_val * 100, 100)
                act_pct = min(abs(actual) / max_val * 100, 100)
                variance = actual - budgeted if d['type'] == 'income' else budgeted - actual
                var_class = 'text-success' if variance >= 0 else 'text-danger'
                var_sign = '+' if variance >= 0 else ''

                rows.append(f'''
                <div class="mb-3">
                    <div class="d-flex justify-content-between mb-1">
                        <strong class="small">{dept_name}</strong>
                        <span class="small {var_class}">
                            {var_sign}${variance:,.0f}
                        </span>
                    </div>
                    <div class="position-relative" style="height: 28px;">
                        <div style="position:absolute;top:0;left:0;height:14px;
                                    width:{bud_pct:.1f}%;background:#a0c4e8;
                                    border-radius:3px;" title="Budgeted: ${budgeted:,.0f}">
                        </div>
                        <div style="position:absolute;top:14px;left:0;height:14px;
                                    width:{act_pct:.1f}%;background:#2e75b6;
                                    border-radius:3px;" title="Actual: ${actual:,.0f}">
                        </div>
                    </div>
                    <div class="d-flex justify-content-between" style="font-size:11px;">
                        <span class="text-muted">
                            <span style="display:inline-block;width:10px;height:10px;
                                         background:#a0c4e8;border-radius:2px;margin-right:3px;">
                            </span>Budget: ${budgeted:,.0f}
                        </span>
                        <span class="text-muted">
                            <span style="display:inline-block;width:10px;height:10px;
                                         background:#2e75b6;border-radius:2px;margin-right:3px;">
                            </span>Actual: ${actual:,.0f}
                        </span>
                    </div>
                </div>''')

            html = (
                '<div class="p-2">'
                '<h5 class="mb-3">'
                '<i class="fa fa-bar-chart text-primary"/> '
                'Budget vs Actual by Department</h5>'
                + ''.join(rows)
                + '</div>'
            )
            rec.department_chart_html = html

    @api.depends("line_ids.amount", "line_ids.actual_amount",
                 "line_ids.account_id", "line_ids.department_id")
    def _compute_pnl_html(self):
        """Build P&L statement HTML grouped by department like QuickBooks."""
        for rec in self:
            if not rec.line_ids:
                rec.pnl_html = (
                    '<div class="text-muted text-center p-4">'
                    '<i class="fa fa-file-text-o fa-2x mb-2"/><br/>'
                    'No budget lines yet.</div>'
                )
                continue

            # Group lines by department, then by type
            dept_data = {}
            for line in rec.line_ids.sorted(key=lambda l: (l.account_code or '')):
                dept = line.department_id
                dept_key = dept.id if dept else 0
                if dept_key not in dept_data:
                    dept_data[dept_key] = {
                        'name': dept.name if dept else 'Unassigned',
                        'code': dept.code if dept else '99',
                        'income': [],
                        'expense': [],
                    }
                entry = {
                    'code': line.account_code or '',
                    'name': line.account_id.name or '',
                    'budget': line.amount or 0.0,
                    'actual': line.actual_amount or 0.0,
                    'variance': line.variance or 0.0,
                    'over': line.is_over_budget,
                }
                if line.account_type == 'income':
                    dept_data[dept_key]['income'].append(entry)
                elif line.account_type in ('expense', 'cogs'):
                    dept_data[dept_key]['expense'].append(entry)

            # Sort departments by code
            sorted_depts = sorted(dept_data.values(), key=lambda d: d['code'])

            # Build HTML
            css = '''
            <style>
                .pnl-table { width:100%; border-collapse:collapse; font-size:13px; }
                .pnl-table th { background:#2E4057; color:#fff; padding:6px 10px;
                                text-align:right; font-weight:600; }
                .pnl-table th:first-child, .pnl-table th:nth-child(2) { text-align:left; }
                .pnl-table td { padding:4px 10px; border-bottom:1px solid #eee; }
                .pnl-table td:nth-child(n+3) { text-align:right; font-family:monospace; }
                .pnl-dept { background:#E8EEF2; font-weight:700; font-size:14px; }
                .pnl-dept td { padding:8px 10px; border-bottom:2px solid #ccc; }
                .pnl-section td { background:#F5F7FA; font-weight:600; font-style:italic;
                                  padding:5px 10px; }
                .pnl-subtotal td { border-top:2px solid #999; font-weight:700; }
                .pnl-total td { background:#2E4057; color:#fff; font-weight:700;
                                font-size:14px; padding:8px 10px; }
                .pnl-neg { color:#dc3545; }
                .pnl-pos { color:#28a745; }
                .pnl-over { background:#FDECEA; }
            </style>
            '''

            def fmt(val):
                if val < 0:
                    return f'<span class="pnl-neg">({abs(val):,.2f})</span>'
                return f'{val:,.2f}'

            def fmt_var(val):
                cls = 'pnl-pos' if val >= 0 else 'pnl-neg'
                return f'<span class="{cls}">{fmt(val)}</span>'

            rows = []
            grand_income_bud = grand_income_act = 0.0
            grand_expense_bud = grand_expense_act = 0.0

            for dept in sorted_depts:
                if not dept['income'] and not dept['expense']:
                    continue

                # Department header
                rows.append(
                    f'<tr class="pnl-dept"><td colspan="5">'
                    f'{dept["name"]}</td></tr>'
                )

                dept_income_bud = dept_income_act = 0.0
                dept_expense_bud = dept_expense_act = 0.0

                # Income section
                if dept['income']:
                    rows.append(
                        '<tr class="pnl-section"><td></td>'
                        '<td>Income</td><td></td><td></td><td></td></tr>'
                    )
                    for item in dept['income']:
                        rows.append(
                            f'<tr><td style="padding-left:30px;color:#666;">'
                            f'{item["code"]}</td>'
                            f'<td style="padding-left:20px;">{item["name"]}</td>'
                            f'<td>{fmt(item["budget"])}</td>'
                            f'<td>{fmt(item["actual"])}</td>'
                            f'<td>{fmt_var(item["variance"])}</td></tr>'
                        )
                        dept_income_bud += item['budget']
                        dept_income_act += item['actual']

                    inc_var = dept_income_act - dept_income_bud
                    rows.append(
                        f'<tr class="pnl-subtotal"><td></td>'
                        f'<td>Total Income</td>'
                        f'<td>{fmt(dept_income_bud)}</td>'
                        f'<td>{fmt(dept_income_act)}</td>'
                        f'<td>{fmt_var(inc_var)}</td></tr>'
                    )

                # Expense section
                if dept['expense']:
                    rows.append(
                        '<tr class="pnl-section"><td></td>'
                        '<td>Expenses</td><td></td><td></td><td></td></tr>'
                    )
                    for item in dept['expense']:
                        over_cls = ' class="pnl-over"' if item['over'] else ''
                        rows.append(
                            f'<tr{over_cls}><td style="padding-left:30px;color:#666;">'
                            f'{item["code"]}</td>'
                            f'<td style="padding-left:20px;">{item["name"]}</td>'
                            f'<td>{fmt(item["budget"])}</td>'
                            f'<td>{fmt(item["actual"])}</td>'
                            f'<td>{fmt_var(item["variance"])}</td></tr>'
                        )
                        dept_expense_bud += item['budget']
                        dept_expense_act += item['actual']

                    exp_var = dept_expense_bud - dept_expense_act
                    rows.append(
                        f'<tr class="pnl-subtotal"><td></td>'
                        f'<td>Total Expenses</td>'
                        f'<td>{fmt(dept_expense_bud)}</td>'
                        f'<td>{fmt(dept_expense_act)}</td>'
                        f'<td>{fmt_var(exp_var)}</td></tr>'
                    )

                # Department net
                dept_net_bud = dept_income_bud - dept_expense_bud
                dept_net_act = dept_income_act - dept_expense_act
                dept_net_var = dept_net_act - dept_net_bud
                net_cls = 'pnl-pos' if dept_net_act >= 0 else 'pnl-neg'
                rows.append(
                    f'<tr style="background:#f0f4f8;"><td></td>'
                    f'<td><strong>Net {dept["name"]}</strong></td>'
                    f'<td><strong>{fmt(dept_net_bud)}</strong></td>'
                    f'<td><strong><span class="{net_cls}">'
                    f'{fmt(dept_net_act)}</span></strong></td>'
                    f'<td><strong>{fmt_var(dept_net_var)}</strong></td></tr>'
                )
                # Spacer row
                rows.append('<tr><td colspan="5" style="height:12px;"></td></tr>')

                grand_income_bud += dept_income_bud
                grand_income_act += dept_income_act
                grand_expense_bud += dept_expense_bud
                grand_expense_act += dept_expense_act

            # Grand totals
            grand_net_bud = grand_income_bud - grand_expense_bud
            grand_net_act = grand_income_act - grand_expense_act
            grand_net_var = grand_net_act - grand_net_bud
            rows.append(
                f'<tr class="pnl-total"><td></td>'
                f'<td>TOTAL INCOME</td>'
                f'<td>{fmt(grand_income_bud)}</td>'
                f'<td>{fmt(grand_income_act)}</td>'
                f'<td>{fmt_var(grand_income_act - grand_income_bud)}</td></tr>'
            )
            rows.append(
                f'<tr class="pnl-total"><td></td>'
                f'<td>TOTAL EXPENSES</td>'
                f'<td>{fmt(grand_expense_bud)}</td>'
                f'<td>{fmt(grand_expense_act)}</td>'
                f'<td>{fmt_var(grand_expense_bud - grand_expense_act)}</td></tr>'
            )
            net_cls = 'pnl-pos' if grand_net_act >= 0 else 'pnl-neg'
            rows.append(
                f'<tr class="pnl-total" style="background:#1a2e3f;">'
                f'<td></td><td>NET INCOME</td>'
                f'<td>{fmt(grand_net_bud)}</td>'
                f'<td><span class="{net_cls}">{fmt(grand_net_act)}</span></td>'
                f'<td>{fmt_var(grand_net_var)}</td></tr>'
            )

            html = (
                f'{css}'
                f'<div class="p-2">'
                f'<table class="pnl-table">'
                f'<thead><tr><th>Code</th><th>Account</th>'
                f'<th>Budget</th><th>Actual</th><th>Variance</th>'
                f'</tr></thead><tbody>'
                + ''.join(rows)
                + '</tbody></table></div>'
            )
            rec.pnl_html = html

    def _validate_budget_for_approval(self):
        """Common validation before any approval step."""
        for rec in self:
            if rec.net_budget < 0:
                raise ValidationError(_(
                    "The overall budget cannot be negative. "
                    "Current net: %.2f"
                ) % rec.net_budget)
            if not rec.line_ids:
                raise ValidationError(_(
                    "Cannot approve an empty budget. Add budget lines first."
                ))

    def action_submit_board(self):
        """Submit budget to the Board for review. Locks the budget."""
        self._validate_budget_for_approval()
        self.write({"state": "board_pending"})
        for rec in self:
            rec.message_post(
                body=_("Submitted to <b>Board</b> for review by %s.",
                       self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_board_approve(self):
        """Board approves the budget → advance to Board Approved."""
        self.write({"state": "board_approved"})
        for rec in self:
            rec.message_post(
                body=_("<b>Board Approved</b> by %s.", self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_board_reject(self):
        """Board rejects the budget."""
        self.write({"state": "rejected"})
        for rec in self:
            rec.message_post(
                body=_("<b>Board Rejected</b> by %s.", self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_submit_floor(self):
        """Submit budget to the Floor for vote."""
        self.write({"state": "floor_pending"})
        for rec in self:
            rec.message_post(
                body=_("Submitted to <b>Floor</b> for vote by %s.",
                       self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_floor_approve(self):
        """Floor approves the budget → Floor Approved."""
        self.write({"state": "floor_approved"})
        for rec in self:
            rec.message_post(
                body=_("<b>Floor Approved</b> by %s.", self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_floor_reject(self):
        """Floor rejects the budget."""
        self.write({"state": "rejected"})
        for rec in self:
            rec.message_post(
                body=_("<b>Floor Rejected</b> by %s.", self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_reset_rejected(self):
        """Reset a rejected budget back to draft."""
        self.write({"state": "draft"})
        for rec in self:
            rec.message_post(
                body=_("Reset to <b>Draft</b> by %s.", self.env.user.name),
                subtype_xmlid='mail.mt_comment',
            )

    def action_view_transfers(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Budget Transfers'),
            'res_model': 'elks.budget.transfer',
            'view_mode': 'list,form',
            'domain': [('budget_id', '=', self.id)],
            'context': {'default_budget_id': self.id},
        }

    def action_view_amendments(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Budget Amendments'),
            'res_model': 'elks.budget.amendment',
            'view_mode': 'list,form',
            'domain': [('budget_id', '=', self.id)],
            'context': {'default_budget_id': self.id},
        }

    def action_generate_csv(self):
        """Generate FRS budget CSV.

        Format: LodgeNumber, LodgeGLAccount, FYE, Version, Annual
        """
        self.ensure_one()
        settings = self.env["elks.lodge.settings"].sudo().search([], limit=1)
        if not settings or not settings.lodge_number:
            raise UserError(_(
                "Please configure your Lodge Number in Elks FRS → Settings."
            ))

        lodge_num = settings.lodge_number
        fye = self.fiscal_year_end.strftime("%m/%d/%Y") if self.fiscal_year_end else ""

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "LodgeNumber", "LodgeGLAccount", "FYE", "Version", "Annual",
        ])
        for line in self.line_ids.sorted(key=lambda l: l.account_id.code):
            acct_code = line.account_id.code
            if line.account_id.subaccount:
                acct_code = f"{acct_code}{line.account_id.subaccount}"
            writer.writerow([
                lodge_num, acct_code, fye,
                self.version or "1",
                f"{line.amount:.2f}",
            ])

        csv_data = output.getvalue()
        filename = f"FRS_Budget_{lodge_num}_{self.lodge_year}.csv"

        self.write({
            "csv_file": base64.b64encode(csv_data.encode("utf-8")),
            "csv_filename": filename,
            "state": "submitted",
        })

    # ------------------------------------------------------------------
    # Print PDF report
    # ------------------------------------------------------------------
    def action_print_budget_report(self):
        """Preview the Budget vs Actual report (HTML)."""
        self.ensure_one()
        return self.env.ref('elksfrs.action_report_budget').report_action(self)

    def action_download_budget_pdf(self):
        """Download the Budget vs Actual report as PDF."""
        self.ensure_one()
        return self.env.ref(
            'elksfrs.action_report_budget_pdf'
        ).report_action(self)

    # ------------------------------------------------------------------
    # Export for reimport (CSV matching import wizard format)
    # ------------------------------------------------------------------
    def action_export_budget_csv(self):
        """Export budget lines to a CSV that the Budget Import Wizard can reimport.

        Format: AccountCode, AccountName, BudgetAmount, Note
        """
        self.ensure_one()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["AccountCode", "AccountName", "BudgetAmount", "Note"])

        for line in self.line_ids.sorted(key=lambda l: l.account_code or ''):
            acct = line.account_id
            code = acct.code or ''
            if acct.subaccount:
                code = f"{code}{acct.subaccount}"
            writer.writerow([
                code,
                acct.name or '',
                f"{line.amount:.2f}",
                line.note or '',
            ])

        csv_bytes = output.getvalue().encode("utf-8")
        filename = f"Budget_Export_{self.lodge_year or 'draft'}.csv"

        # Store on a transient attachment so the user gets a download
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(csv_bytes),
            'mimetype': 'text/csv',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # Excel export with Budget vs Actuals
    # ------------------------------------------------------------------
    def action_export_budget_xlsx(self):
        """Export budget with actuals to a formatted Excel workbook."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        except ImportError:
            raise UserError(_(
                "The openpyxl library is required for Excel export. "
                "Please install it: pip install openpyxl"
            ))

        wb = openpyxl.Workbook()

        # ----- Style definitions -----
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2E4057')
        title_font = Font(name='Arial', bold=True, size=14)
        subtitle_font = Font(name='Arial', bold=True, size=11, color='555555')
        money_fmt = '$#,##0.00;($#,##0.00);"-"'
        pct_fmt = '0.0%'
        thin_border = Border(
            bottom=Side(style='thin', color='CCCCCC'),
        )
        total_font = Font(name='Arial', bold=True, size=11)
        total_fill = PatternFill('solid', fgColor='E8EEF2')
        over_budget_fill = PatternFill('solid', fgColor='FDECEA')
        good_fill = PatternFill('solid', fgColor='E8F5E9')

        def _write_section(ws, lines, section_title, start_row):
            """Write a budget section (Income or Expense) and return next row."""
            row = start_row

            # Section header
            ws.cell(row=row, column=1, value=section_title).font = subtitle_font
            row += 1

            # Column headers
            headers = ['Account Code', 'Account Name', 'Budgeted', 'Actual',
                        'Variance', '% Used']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal='center')
            row += 1

            # Data rows
            total_budgeted = 0
            total_actual = 0
            total_variance = 0

            for line in lines.sorted(key=lambda l: l.account_code or ''):
                acct = line.account_id
                code = acct.code or ''
                if acct.subaccount:
                    code = f"{code}{acct.subaccount}"

                ws.cell(row=row, column=1, value=code).font = Font(name='Arial', size=10)
                ws.cell(row=row, column=2, value=acct.name or '').font = Font(name='Arial', size=10)

                c_bud = ws.cell(row=row, column=3, value=line.amount)
                c_bud.number_format = money_fmt
                c_act = ws.cell(row=row, column=4, value=line.actual_amount)
                c_act.number_format = money_fmt
                c_var = ws.cell(row=row, column=5, value=line.variance)
                c_var.number_format = money_fmt
                c_pct = ws.cell(row=row, column=6,
                                value=line.percent_used / 100.0 if line.percent_used else 0)
                c_pct.number_format = pct_fmt

                # Highlight over-budget rows
                if line.is_over_budget:
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).fill = over_budget_fill
                elif line.variance > 0:
                    c_var.fill = good_fill

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border

                total_budgeted += line.amount or 0
                total_actual += line.actual_amount or 0
                total_variance += line.variance or 0
                row += 1

            # Totals row
            ws.cell(row=row, column=2, value='TOTAL').font = total_font
            c = ws.cell(row=row, column=3, value=total_budgeted)
            c.number_format = money_fmt
            c.font = total_font
            c.fill = total_fill
            c = ws.cell(row=row, column=4, value=total_actual)
            c.number_format = money_fmt
            c.font = total_font
            c.fill = total_fill
            c = ws.cell(row=row, column=5, value=total_variance)
            c.number_format = money_fmt
            c.font = total_font
            c.fill = total_fill

            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = total_fill

            return row + 2  # skip a blank row

        # ----- Build the workbook -----
        ws = wb.active
        ws.title = "Budget vs Actual"

        # Page setup for printing
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Title
        ws.cell(row=1, column=1, value=self.name).font = title_font
        ws.cell(row=2, column=1, value=f"Lodge Year: {self.lodge_year or ''}").font = subtitle_font
        ws.cell(row=2, column=4,
                value=f"FYE: {self.fiscal_year_end.strftime('%m/%d/%Y') if self.fiscal_year_end else ''}").font = subtitle_font

        row = 4

        # Income section
        income_lines = self.line_ids.filtered(lambda l: l.account_type == 'income')
        if income_lines:
            row = _write_section(ws, income_lines, "INCOME", row)

        # Expense section
        expense_lines = self.line_ids.filtered(lambda l: l.account_type in ('expense', 'cogs'))
        if expense_lines:
            row = _write_section(ws, expense_lines, "EXPENSES", row)

        # Net summary
        ws.cell(row=row, column=2, value='NET (Income − Expenses)').font = Font(
            name='Arial', bold=True, size=12)
        c = ws.cell(row=row, column=3, value=self.net_budget)
        c.number_format = money_fmt
        c.font = Font(name='Arial', bold=True, size=12)
        c = ws.cell(row=row, column=4, value=self.actual_net)
        c.number_format = money_fmt
        c.font = Font(name='Arial', bold=True, size=12)

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 12

        # Print headers
        ws.print_title_rows = '1:3'
        ws.oddHeader.center.text = self.name or "Budget Report"
        ws.oddFooter.left.text = "Printed: &D"
        ws.oddFooter.right.text = "Page &P of &N"

        # Save to binary
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_data = buf.getvalue()
        filename = f"Budget_{self.lodge_year or 'draft'}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_reset_draft(self):
        self.write({
            "state": "draft",
            "csv_file": False,
            "csv_filename": False,
        })

    # ------------------------------------------------------------------
    # Update Dues Income from paid members (manual catch-up button)
    # ------------------------------------------------------------------
    def action_update_dues_income(self):
        """Manual catch-up: count any paid members not yet counted.

        Normally dues income is created automatically when the
        x_is_dues_paid flag flips True (see res_partner_frs.py).
        This button handles edge cases:
        - First-time setup after importing contacts
        - Members who slipped through before the auto-trigger existed
        - Manual re-sync if something went wrong
        """
        self.ensure_one()
        if not self.fiscal_year_end:
            raise UserError(_("Set the Fiscal Year End before updating dues income."))

        lodge_year = self.lodge_year
        if not lodge_year:
            start_yr = self.fiscal_year_end.year - 1
            lodge_year = f"{start_yr}-{self.fiscal_year_end.year}"

        Partner = self.env['res.partner']

        # Find uncounted paid members
        uncounted = Partner.search([
            ('x_is_member', '=', True),
            ('x_is_dues_paid', '=', True),
            '|',
            ('x_dues_budget_year', '=', False),
            ('x_dues_budget_year', '!=', lodge_year),
        ])

        if not uncounted:
            raise UserError(_(
                "No new paid members to count.\n\n"
                "All members with current dues have already been "
                "counted for lodge year %s."
            ) % lodge_year)

        # Delegate to the same auto-count logic on res.partner
        uncounted._auto_count_dues_for_budget()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Dues Income Updated'),
                'message': _(
                    "Counted %d paid members for lodge year %s. "
                    "Check the budget chatter for details."
                ) % (len(uncounted), lodge_year),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    # ------------------------------------------------------------------
    # Dashboard navigation helpers
    # ------------------------------------------------------------------
    def action_view_budget_detail(self):
        """Open the full budget form view."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Budget Detail'),
            'res_model': 'elks.budget',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_view_budget_vs_actual(self):
        """Open budget lines in Budget vs Actual list, grouped by type."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Budget vs Actual — %s') % self.name,
            'res_model': 'elks.budget.line',
            'view_mode': 'list',
            'domain': [('budget_id', '=', self.id)],
            'context': {
                'search_default_group_type': 1,
            },
        }

    @api.model
    def action_open_dashboard(self):
        """Server action: find current lodge year budget and open dashboard.

        Used as the FRS Dashboard menu entry point so the user lands on
        the right budget automatically.
        """
        today = fields.Date.context_today(self)
        # Lodge fiscal year ends March 31.  If today is Jan–Mar we are in
        # the year that ends this March; otherwise it ends next March.
        if today.month <= 3:
            fye = today.replace(month=3, day=31)
        else:
            fye = today.replace(year=today.year + 1, month=3, day=31)

        budget = self.search([('fiscal_year_end', '=', fye)], limit=1)
        if not budget:
            # Fall back to the most recent budget
            budget = self.search([], order='fiscal_year_end desc', limit=1)

        if not budget:
            # No budget yet — open the budget list so the user can create one
            return {
                'type': 'ir.actions.act_window',
                'name': _('Budgets'),
                'res_model': 'elks.budget',
                'view_mode': 'list,form',
                'target': 'current',
            }

        return {
            'type': 'ir.actions.act_window',
            'name': _('FRS Dashboard'),
            'res_model': 'elks.budget',
            'res_id': budget.id,
            'view_mode': 'form',
            'view_id': self.env.ref('elksfrs.view_elks_budget_dashboard_form').id,
            'target': 'current',
        }

    def _migrate_old_approved_state(self):
        """One-time migration: rename old 'approved' state to 'board_approved'.

        Call via shell or let init_hook run it after module upgrade.
        """
        self.env.cr.execute("""
            UPDATE elks_budget
            SET state = 'board_approved'
            WHERE state = 'approved'
        """)
        count = self.env.cr.rowcount
        if count:
            _logger.info("Migrated %d budgets from 'approved' → 'board_approved'", count)


class ElksBudgetLine(models.Model):
    _name = "elks.budget.line"
    _description = "Elks Budget Line"
    _order = "account_id"

    budget_id = fields.Many2one(
        "elks.budget", string="Budget", required=True,
        ondelete="cascade", index=True,
    )
    account_id = fields.Many2one(
        "elks.account", string="Account", required=True,
        domain="[('is_header', '=', False)]",
    )
    account_code = fields.Char(
        related="account_id.code", store=True, string="Code",
    )
    account_type = fields.Selection(
        related="account_id.account_type", store=True, string="Type",
    )
    department_id = fields.Many2one(
        related="account_id.department_id", store=True,
    )
    amount = fields.Monetary(
        "Budgeted", currency_field='currency_id',
        help="Budgeted annual amount for this account.",
    )
    currency_id = fields.Many2one(
        "res.currency",
        related="budget_id.currency_id",
        store=True,
        default=lambda self: self.env.company.currency_id,
    )
    note = fields.Char("Note")

    # Odoo 19 replacement for _sql_constraints
    @api.constrains("budget_id", "account_id")
    def _check_unique_account_per_budget(self):
        for rec in self:
            if not rec.account_id:
                continue
            dupes = self.search([
                ('budget_id', '=', rec.budget_id.id),
                ('account_id', '=', rec.account_id.id),
                ('id', '!=', rec.id),
            ])
            if dupes:
                raise ValidationError(_(
                    "Account %(acct)s already has a budget line on "
                    "%(budget)s.  Each account can appear only once per budget."
                ) % {
                    'acct': rec.account_id.display_name or rec.account_id.code,
                    'budget': rec.budget_id.name,
                })

    def action_view_actual_journal_lines(self):
        """Open the journal entry lines that make up this line's actual."""
        import datetime
        self.ensure_one()
        if not self.account_id or not self.budget_id.fiscal_year_end:
            return False
        fye = self.budget_id.fiscal_year_end
        fy_start = datetime.date(fye.year - 1, 4, 1)
        return {
            'type': 'ir.actions.act_window',
            'name': _('%(acct)s Activity — FY %(yr)s', acct=self.account_id.code, yr=self.budget_id.lodge_year),
            'res_model': 'elks.journal.entry.line',
            'view_mode': 'list,form',
            'domain': [
                ('account_id', '=', self.account_id.id),
                ('entry_state', '=', 'posted'),
                ('date', '>=', fy_start),
                ('date', '<=', fye),
            ],
            'context': {'search_default_group_date': 1},
        }

    # --- Encumbrance (from approved purchase orders) ---
    encumbered_amount = fields.Monetary(
        "Encumbered", compute="_compute_encumbered", store=False,
        currency_field='currency_id',
        help="Funds reserved by approved Purchase Orders not yet fully spent.",
    )
    available_amount = fields.Monetary(
        "Available", compute="_compute_encumbered", store=False,
        currency_field='currency_id',
        help="Budget remaining after actuals and encumbrances.",
    )

    # --- Actuals (computed from elks.journal.entry.line) ---
    actual_amount = fields.Monetary(
        "Actual", compute="_compute_actuals", store=False,
        currency_field='currency_id',
        help="Sum of posted journal entry lines for this account within "
             "the budget's fiscal year.",
    )
    variance = fields.Monetary(
        "Variance", compute="_compute_actuals", store=False,
        currency_field='currency_id',
        help="For income accounts: actual − budgeted (positive is good). "
             "For expense accounts: budgeted − actual (positive is good).",
    )
    percent_used = fields.Float(
        "% of Budget", compute="_compute_actuals", store=False,
        help="Actual as a percentage of budgeted amount.",
    )
    is_over_budget = fields.Boolean(
        "Over Budget", compute="_compute_actuals", store=False,
    )

    def action_open_detail(self):
        """Open this budget line in a popup form."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _("Budget Line: %s") % self.account_id.display_name,
            'res_model': 'elks.budget.line',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    @api.depends(
        "amount", "account_id", "account_id.account_type",
        "budget_id.fiscal_year_end",
    )
    def _compute_actuals(self):
        """Sum posted journal entry lines against this account within
        the budget's fiscal year (April 1 – March 31)."""
        import datetime
        JELine = self.env['elks.journal.entry.line']

        for line in self:
            if not line.account_id or not line.budget_id.fiscal_year_end:
                line.actual_amount = 0.0
                line.variance = 0.0
                line.percent_used = 0.0
                line.is_over_budget = False
                continue

            fye = line.budget_id.fiscal_year_end
            fy_start = datetime.date(fye.year - 1, 4, 1)

            lines = JELine.search([
                ('account_id', '=', line.account_id.id),
                ('entry_state', '=', 'posted'),
                ('date', '>=', fy_start),
                ('date', '<=', fye),
            ])

            # Income accounts: credits increase income, so actual = credits − debits
            # Expense/COGS accounts: debits increase expense, so actual = debits − credits
            credits = sum(lines.mapped('credit'))
            debits = sum(lines.mapped('debit'))
            acct_type = line.account_id.account_type

            if acct_type == 'income':
                actual = credits - debits
                # Variance: positive = earning more than budgeted (good)
                variance = actual - line.amount
                over = False  # income "over" is good, not a warning
            elif acct_type in ('expense', 'cogs'):
                actual = debits - credits
                # Variance: positive = spent less than budgeted (good)
                variance = line.amount - actual
                over = actual > line.amount
            else:
                actual = debits - credits
                variance = line.amount - actual
                over = False

            line.actual_amount = actual
            line.variance = variance
            # percentage widget expects a fraction (0.0–1.0); it multiplies
            # by 100 for display automatically
            line.percent_used = (
                (actual / line.amount) if line.amount else 0.0
            )
            line.is_over_budget = over

    @api.depends("amount", "budget_id")
    def _compute_encumbered(self):
        """Sum approved PO line amounts charged to this budget line's account.

        Looks at individual PO lines (purchase.order.line) where the
        x_elks_account_id matches this budget line's account, and the
        parent PO has been floor-approved (x_approval_state = 'approved').

        Gracefully handles the case where elkspurchase is not installed.
        """
        POLine = self.env.get('purchase.order.line')
        has_line_account = (
            POLine is not None
            and 'x_elks_account_id' in POLine._fields
        )

        # Check if the x_ordered field exists (elkspurchase may add it)
        has_ordered = (
            has_line_account
            and 'x_ordered' in POLine._fields
        )

        for line in self:
            if has_line_account and line.account_id:
                domain = [
                    ('x_elks_account_id', '=', line.account_id.id),
                    ('order_id.x_approval_state', '=', 'approved'),
                ]
                # Exclude lines already marked as ordered — those have
                # journal entries and show up in actuals instead.
                if has_ordered:
                    domain.append(('x_ordered', '=', False))
                po_lines = POLine.search(domain)
                line.encumbered_amount = sum(
                    po_lines.mapped('price_subtotal')
                )
            else:
                line.encumbered_amount = 0.0
            line.available_amount = (
                line.amount - line.actual_amount - line.encumbered_amount
            )
